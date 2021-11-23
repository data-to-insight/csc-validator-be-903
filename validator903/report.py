import logging
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from validator903.config import configured_errors

logger = logging.getLogger(__name__)

_errors_df = pd.DataFrame([
    (c[0].code, c[0].description, c[0].affected_fields, c[0].sortable_code) for c in configured_errors
], columns=['Code', 'Description', 'Fields', 'Sortable']).sort_values('Sortable')
logger.debug(f"Created error summary df with {_errors_df.shape[0]} configured errors.")


class Report:
    __child_report = None
    __child_summary = None
    __error_report = None
    __error_summary = None

    def __init__(self, data_store):
        dataframes = []
        for table, value in data_store.items():
            if table == "metadata":
                continue
            cols_error = [c for c in value.columns if c[:4] == "ERR_"]
            cols_data = [c for c in value.columns if c[:4] != "ERR_" and c[:2] != "__" and c != "CHILD"]

            df_error = value[['CHILD'] + cols_error].copy()
            df_error['Table'] = table
            df_error['RowID'] = df_error.index
            df_error['Context'] = value[cols_data].apply(lambda row: {**row}, axis=1)

            dataframes.append(df_error)

        report = pd.concat(dataframes, ignore_index=True)
        self.child_cols = ['Table', 'RowID', 'CHILD', 'Context']
        self.error_cols = [c for c in report.columns if c not in self.child_cols]
        report['Total Errors'] = report[self.error_cols].sum(axis=1)
        self.report = report[self.child_cols + self.error_cols + ['Total Errors']]

    @property
    def error_report(self):
        """
        The error report is a simple report containing the number of errors report for each error code
        as well as the error description and the affected fields.
        """
        if self.__error_report is None:
            df = pd.DataFrame(self.report[self.error_cols].sum(), columns=['Count']).reset_index()
            df['Code'] = df['index'].str[4:]
            df = df[['Code', 'Count']]
            self.__error_report = df.merge(_errors_df, on='Code', how='left').sort_values('Sortable')[[
                'Code', 'Description', 'Fields', 'Count',
            ]]
        return self.__error_report

    @property
    def error_summary(self):
        if self.__error_summary is None:
            headers = ["Code", "Error Count", "Error Description", "Affected Fields"]
            df = self.error_report.rename(columns={
                'Description': 'Error Description',
                'Count': 'Error Count',
            })
            df['Affected Fields'] = df['Fields'].apply(lambda x: ", ".join(x))
            self.__error_summary = df[headers]

        return self.__error_summary

    @property
    def child_report(self):
        """
        The child report contains details of every error reported, including context information with the data
        from the affected row.
        """
        if self.__child_report is None:
            child_report = self.report.melt(id_vars=self.child_cols, value_vars=self.error_cols,
                                            var_name='ErrorColumn', value_name='IsError')
            child_report = child_report[child_report.IsError == True].reset_index()
            child_report['Code'] = child_report.ErrorColumn.str[4:]

            child_error_count = child_report[['CHILD', 'Code']].groupby('CHILD').agg(
                **{'Child Error Count': ('Code', 'count')}
            ).reset_index()

            child_report = child_report.merge(child_error_count, on='CHILD', how='left')

            error_report = self.error_report.rename(columns=dict(Count='Error Type Count'))
            child_report = child_report.merge(error_report, on=['Code'], how='left')
            child_report = child_report[['CHILD', 'Table', 'RowID', 'Context', 'Code',
                                         'Description', 'Fields', 'Child Error Count', 'Error Type Count']]
            self.__child_report = child_report
        return self.__child_report

    @property
    def child_summary(self):
        if self.__child_summary is None:
            self.__child_summary = _create_child_summary(self.child_report)
        return self.__child_summary

    def excel_report(self, file=None):
        """
        Creates an excel report containing an error summary and full details of each error with context information.

        If a string or file-like is provided, then the output is saved to this,
        otherwise the workbook data are returned as a buffer.
        """
        logger.info("Creating workbook")
        wb = Workbook(write_only=True)

        error_sheet = wb.create_sheet(title="Error Summary")
        _populate_error_sheet(error_sheet, self.error_summary)

        child_sheet = wb.create_sheet(title="Child Summary")
        _populate_child_sheet(child_sheet, self.child_summary)

        logger.info("Creating workbook - saving")
        if file:
            wb.save(file)
        else:
            io = BytesIO()
            wb.save(io)
            return io.getbuffer()

    def csv_report(self, sheetname, file=None):
        if sheetname == "errors":
            df = self.error_summary
        elif sheetname == "children":
            df = self.child_summary
        else:
            raise ValueError("Unknown export type requested")

        if file:
            df.to_csv(file, index=False)
        else:
            return df.to_csv(index=False)


def _populate_error_sheet(error_sheet, df):
    logger.debug("Creating workbook - error summary")

    error_sheet.column_dimensions["A"].width = 15
    error_sheet.column_dimensions["B"].width = 15
    error_sheet.column_dimensions["C"].width = 60
    error_sheet.column_dimensions["D"].width = 120

    headers = [c for c in df.columns]
    error_sheet.append(headers)
    for r in dataframe_to_rows(df, index=False, header=False):
        error_sheet.append(r)

    error_table = Table(
        displayName="Errors",
        ref=f"A1:{get_column_letter(len(headers))}{df.shape[0] + 1}"
    )
    error_table._initialise_columns()
    for column, value in zip(error_table.tableColumns, headers):
        column.name = value

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    error_table.tableStyleInfo = style
    error_sheet.add_table(error_table)


def _populate_child_sheet(child_sheet, df):
    logger.debug("Creating workbook - child summary")

    child_sheet.column_dimensions["A"].width = 15
    child_sheet.column_dimensions["B"].width = 15
    child_sheet.column_dimensions["C"].width = 15
    child_sheet.column_dimensions["D"].width = 15
    child_sheet.column_dimensions["E"].width = 60
    child_sheet.column_dimensions["F"].width = 30
    child_sheet.column_dimensions["G"].width = 40
    child_sheet.column_dimensions["H"].width = 20
    child_sheet.column_dimensions["I"].width = 15
    child_sheet.column_dimensions["J"].width = 15

    logger.debug("Child Summary - Adding headers")
    headers = [c for c in df.columns]
    child_sheet.append(headers)

    logger.debug("Child Summary - Populating rows")
    for index, row in df.iterrows():
        child_sheet.append([row[h] for h in headers])

    logger.debug("Child Summary - Adding table")
    child_table = Table(
        displayName="ChildSummary",
        ref=f"A1:{get_column_letter(len(headers))}{df.shape[0] + 1}"
    )
    child_table._initialise_columns()
    for column, value in zip(child_table.tableColumns, headers):
        column.name = value

    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    child_table.tableStyleInfo = style
    child_sheet.add_table(child_table)


def _create_child_summary(child_report: pd.DataFrame) -> pd.DataFrame:
    logger.debug("Child Summary - Renaming Columns")
    headers = [
        "Child",
        "Affected Table",
        "Error Row",
        "Error Code",
        "Error Description",
        "Error Fields",
        "Affected Values",
        "Locator Hints",
        "Errors for Child",
        "Errors of Type",
    ]
    df = child_report.rename(columns={
        'CHILD': 'Child',
        'Table': 'Affected Table',
        'RowID': 'Error Row',
        'Code': 'Error Code',
        'Description': 'Error Description',
        'Child Error Count': "Errors for Child",
        'Error Type Count': "Errors of Type",
    })
    logger.debug("Child Summary - Setting Error Fields")
    df['Error Fields'] = df['Fields'].apply(lambda x: ", ".join(x))
    logger.debug("Child Summary - Setting Affected Values")
    df['Affected Values'] = df[['Fields', 'Context']].apply(
        lambda row: ", ".join([f"{f}: {row.Context[f]}" for f in row.Fields if f in row.Context]),
        axis=1
    )
    logger.debug("Child Summary - Setting Locator Hints")
    df['Locator Hints'] = df.Context.apply(
        lambda row: ", ".join(
            [f"{f}: {row[f]}" for f in ["DECOM", "MIS_START", "REVIEW"] if f in row]
        )
    )
    logger.debug("Child Summary - Sorting")
    df.sort_values(['Child', 'Affected Table', 'Error Code'], inplace=True)
    return df[headers]
